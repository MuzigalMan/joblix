from scraper import scraper as sc
import pandas as pd
from reparser import parser
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Color
from openpyxl.styles.colors import BLUE



def getData(job,location):
    scraper = sc(job,location)
    data = scraper.linkedin_data_1()

def main(data) : 
    #get data
    # get_job = list(data.items())[0]
    job = data['job']
    # get_location = list(data.items())[1]
    location = data['location']

    #resume parser
    user = parser()
    userExp = user['exp']
    userSkill = user['skills']

    if userExp in (1,2,3):
        userLevel = ["Entry level","Internship","Associate"]
    elif userExp >= 4 and userExp <= 10:
        userLevel = ["Mid-Senior level","Executive"]
    elif userExp > 10:
        userLevel = "Director"

    data = pd.read_csv('./data/job_data')
    data['Date'] = pd.to_datetime(data['Date'])

    if isinstance(userLevel, str):
        df_filtered = data[data['Experience'] == userLevel]
    elif isinstance(userLevel, list):
        df_filtered = data[data['Experience'].isin(userLevel)]
    else:
        raise ValueError('userLevel must be a string or a list')
    
    # scraper
    getData(job,location)

    df_filtered['Matched Skills'] = df_filtered['JD'].apply(lambda x: [s for s in userSkill if s.lower() in x.lower()])

    df_filtered = df_filtered[df_filtered['Matched Skills'].apply(len) >= 5]

    df_filtered = df_filtered.sort_values('Date', ascending=False)

    df_filtered['Matched Skills'] = df_filtered['Matched Skills'].apply(lambda x: ', '.join(x))

    wb = Workbook()
    ws = wb.active
    ws.title = f'{job} Jobs'

    headers = df_filtered.columns.tolist()
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=header)

    for r, row in enumerate(df_filtered.itertuples(index=False), start=2):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=value)
            if cell.column == get_column_letter(headers.index('Link') + 1):
                cell.font = Font(color=BLUE, underline='single')
                cell.hyperlink = cell.value

    for c in range(1, len(headers)+1):
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].auto_size = True

    wb.save('result/Datalix.xlsx')

    # return job


# getData("Full Stack Developer","Hyderabad") 
# temp = {'job': 'watchman', 'location':'hyd'}
# main(temp)